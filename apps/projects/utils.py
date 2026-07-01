from io import BytesIO

from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl import Workbook

from django.db import transaction

from rest_framework import serializers

from apps.projects.models import Task


def export_tasks(task_type):
    wb = Workbook()
    ws = wb.active
    ws.title = task_type.name

    ws.append(['Task list'])

    queryset = (
        Task.objects
        .filter(
            task_type=task_type,
            is_active=True,
        )
        .order_by('name')
    )

    for task in queryset:
        ws.append([task.name])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

@transaction.atomic
def import_tasks(task_type, file):
    try:
        file.seek(0)
        workbook = load_workbook(file)
    except BadZipFile:
        raise serializers.ValidationError({
            'file': 'Uploaded file is not a valid Excel (.xlsx) file.'
        })
    except Exception:
        raise serializers.ValidationError({
            'file': 'Unable to read the uploaded Excel file.'
        })

    sheet = workbook.active

    excel_tasks = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        task_name = ' '.join(str(row[0]).split())

        excel_tasks.add(task_name)

        Task.objects.update_or_create(
            task_type=task_type,
            name=task_name,
            defaults={
                'is_active': True,
            }
        )

    (
        Task.objects
        .filter(task_type=task_type)
        .exclude(name__in=excel_tasks)
        .update(is_active=False)
    )
