from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from django.http import HttpResponse


HEADER_FILL = PatternFill(
    fill_type='solid',
    fgColor='D9D9D9',
)

HEADER_FONT = Font(bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

CENTER = Alignment(
    horizontal='center',
    vertical='center',
)

def get_headers(report_days):
    return [
        'id',
        'full_name',
        'department',
        'position',
        'grade',

        *[str(day.day) for day in report_days],

        'working_days',
        'worked_days',
        'worked_hours',
        'weekends_holidays',
        'paid_leaves',
        'non_paid_leaves',
        'business_trips',
        'sick_leaves',
        'day_offs',
        'maternity_leave_days',
        'skipped_days',
    ]

def write_headers(ws, headers):
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def write_row(ws, headers, row):
    values = [row.get(header, '') for header in headers]

    ws.append(values)

    current_row = ws.max_row

    for cell in ws[current_row]:
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def autosize_columns(ws):
    for column in ws.columns:
        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:
            value = '' if cell.value is None else str(cell.value)

            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = max_length + 2

def workbook_to_response(workbook):
    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type=(
            'application/'
            'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )

    response['Content-Disposition'] = (
        'attachment; filename="attendance.xlsx"'
    )

    return response

def export_attendance_excel(report, report_days):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Attendance'
    rows = report['report']

    if not rows:
        worksheet.append(['No data'])
        return workbook_to_response(workbook)

    headers = get_headers(report_days)

    write_headers(worksheet, headers)

    for row in rows:
        write_row(worksheet, headers, row)

    autosize_columns(worksheet)

    return workbook_to_response(workbook)
