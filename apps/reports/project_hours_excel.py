from io import BytesIO

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(
    fill_type='solid',
    fgColor='D9D9D9',
)

HEADER_FONT = Font(bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

CENTER = Alignment(
    horizontal='center',
    vertical='center',
)


def get_headers(start_date, end_date):
    days = range(
        start_date.day,
        end_date.day + 1,
    )

    return [
        'user_id',
        'full_name',
        'department',
        'position',
        'grade',
        'project_id',
        'project_code',

        *[str(day) for day in days],

        'total_hours',
    ]


def export_project_hours_excel(
    report,
    start_date,
    end_date,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Project Hours'

    headers = get_headers(
        start_date,
        end_date,
    )

    write_headers(worksheet, headers)

    for row in report:
        write_row(
            worksheet,
            headers,
            row,
        )

    autosize_columns(worksheet)

    return workbook_to_response(workbook)


def write_headers(ws, headers):
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER


def write_row(ws, headers, row):
    values = [
        row.get(header, '')
        for header in headers
    ]

    ws.append(values)

    for cell in ws[ws.max_row]:
        cell.border = THIN_BORDER
        cell.alignment = CENTER


def autosize_columns(ws):
    for column in ws.columns:
        max_length = max(
            len(str(cell.value or ''))
            for cell in column
        )

        column_letter = get_column_letter(
            column[0].column
        )

        ws.column_dimensions[column_letter].width = (
            max_length + 2
        )


def workbook_to_response(workbook):
    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type=(
            'application/'
            'vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )

    response['Content-Disposition'] = (
        'attachment; '
        'filename="project_hours.xlsx"'
    )

    return response