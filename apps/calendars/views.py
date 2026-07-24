from datetime import datetime

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl

from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from django.db.models import F
from django.http import HttpResponse

from apps.calendars.pagination import TimeEntryPagination
from apps.reports.attendance_excel import export_attendance_excel
from apps.reports.attendance_report import build_report

from .utils import get_country, filter_report_by_params, filter_leaves_by_params
from .services.monitoring import get_monitoring_data
from .services.dashboard import get_dashboard_data
from .permissions import IsOwnerOrAdmin
from .serializers import *
from .models import *


class CountrySettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        country, error = get_country(request)

        if error:
            return error

        settings = CountrySettings.get_settings(country.id)
        serializer = CountrySettingsSerializer(settings)

        return Response(serializer.data)

    def patch(self, request):
        country, error = get_country(request)

        if error:
            return error

        settings = CountrySettings.get_settings(country.id)
        serializer = CountrySettingsSerializer(
            settings,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request):
        country, error = get_country(request)

        if error:
            return error

        settings = CountrySettings.get_settings(country.id)
        serializer = CountrySettingsSerializer(settings, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TimeEntryViewSet(ModelViewSet):
    queryset = TimeEntry.objects.all().select_related(
        'user',
        'country',
        'client',
        'project_code',
        'task_type',
        'task'
    )
    serializer_class = TimeEntryReadSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    pagination_class = TimeEntryPagination

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            return TimeEntryReadSerializer

        return TimeEntryWriteSerializer

    def get_queryset(self):
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        base_queryset = super().get_queryset()
        user = self.request.user

        if start_date and end_date:
            base_queryset = base_queryset.filter(date__gte=start_date, date__lte=end_date)

        return base_queryset.filter(user=user.id)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def get_serializer_context(self):
        return {'request': self.request}

    def paginate_queryset(self, queryset):
        if self.action == 'list' or self.action == 'dashboard':
            return None

        return super().paginate_queryset(queryset)

    def _generate_excel(self, queryset):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Timesheet Report'

        columns = [
            'date',
            'user_email',
            'country_code',
            'user_department',
            'position',
            'detailed_grade',
            'hours',
            'project_department',
            'client_name',
            'code',
            'project_service_line',
            'task_type_name',
            'task_name',
            'description',
        ]

        headers = [
            'Date',
            'User Email',
            'Country',
            'User Department',
            'Position',
            'Grade',
            'Hours',
            'Project Department',
            'Client',
            'Project Code',
            'Service Line',
            'Task Type',
            'Task',
            'Description',
        ]

        sheet.append(headers)

        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).font = Font(bold=True)

        for row in queryset:
            sheet.append([row.get(col) for col in columns])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=timesheet_report.xlsx'

        workbook.save(response)

        return response

    @action(detail=False, methods=['get'])
    def report(self, request):
        queryset = TimeEntry.objects.all().select_related(
            'user',
            'country',
            'client',
            'project_code',
            'task_type',
            'task'
        )

        queryset = filter_report_by_params(request, queryset)

        queryset = (
            queryset.annotate(
                user_email=F('user__email'),
                user_country_code=F('user__country__code'),
                country_code=F('country__code'),
                user_department=F('user__department__name'),
                position=F('user__position__name'),
                detailed_grade=F('user__grade__name'),
                client_name=F('client__name'),
                project_department=F('project_code__project__department__name'),
                code=F('project_code__code'),
                project_service_line=F('project_code__project__service_line__name'),
                task_type_name=F('task_type__name'),
                task_name=F('task__name'),
                status_name=F('user__status__name')
            )
            .values(
                'date',
                'user_email',
                'user_country_code',
                'country_code',
                'user_department',
                'position',
                'detailed_grade',
                'hours',
                'project_department',
                'client_name',
                'code',
                'project_service_line',
                'task_type_name',
                'task_name',
                'description',
                'status_name',
            )
            .order_by('user_email', 'date')
        )

        format_type = request.query_params.get('export', 'json')

        if format_type == 'excel':
            return self._generate_excel(queryset)

        page = self.paginate_queryset(queryset)

        return self.get_paginated_response(page)

    @action(detail=False, methods=['get'])
    def leaves(self, request):
        queryset = (
            self.queryset
            .filter(task_type__name='Leave')
            .order_by('user__email', 'date')
        )

        queryset = filter_leaves_by_params(request, queryset)

        # format_type = request.query_params.get('export', 'json')

        # if format_type == 'excel':
        #     return self._generate_excel(queryset)

        page = self.paginate_queryset(queryset)

        serializer = LeaveReportSerializer(
            page,
            many=True,
        )

        return self.get_paginated_response(serializer.data)

    @action(detail=False, methods=['get'])
    def monitoring(self, request):
        start_date = datetime.strptime(request.query_params.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.query_params.get('end_date'), '%Y-%m-%d').date()
        country_id = request.query_params.get('country_id')

        data = get_monitoring_data(request, country_id, start_date, end_date)

        page = self.paginate_queryset(data)

        return self.get_paginated_response(page)

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        month = int(request.query_params.get('month'))
        year = int(request.query_params.get('year'))
        user = request.user

        data = get_dashboard_data(user, year, month)

        return Response(data)

    @action(detail=False, methods=['get'])
    def attendance(self, request):
        serializer = AttendanceReportSerializer(
            data=request.query_params
        )
        serializer.is_valid(raise_exception=True)

        start_date = serializer.validated_data['start_date']
        end_date = serializer.validated_data['end_date']

        report_days = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]

        report = build_report(
            country_id=serializer.validated_data['country_id'],
            report_start=start_date,
            report_end=end_date,
        )

        format_type = request.query_params.get('export', 'json')

        if format_type == 'excel':
            return export_attendance_excel(report, report_days)

        return Response(report['report'])


class CalendarViewSet(ModelViewSet):
    queryset = Calendar.objects.all().select_related(
        'country',
    )
    serializer_class = CalendarSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    @action(detail=False, methods=['get'], url_path='day-types')
    def day_types(self, request):
        return Response({
            'day_types': [
                {
                    'value': value, 
                    'label': label
                }
                for value, label in Calendar.DayType.choices
            ]
        })

    @action(detail=False, methods=['get'])
    def holidays(self, request):
        user = request.user

        filtered_queryset = Calendar.objects.filter(day_type=Calendar.DayType.HOLIDAY, country=user.country.id)

        serializer = self.get_serializer(filtered_queryset, many=True)

        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='working-weekends')
    def working_weekends(self, request):
        user = request.user

        filtered_queryset = Calendar.objects.filter(day_type=Calendar.DayType.WORKING_WEEKEND, country=user.country.id)

        serializer = self.get_serializer(filtered_queryset, many=True)

        return Response(serializer.data)
