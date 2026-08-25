from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl

from django.http import HttpResponse

from apps.accounts.models import UserStatus


def is_admin(user):
    return (
        user.is_authenticated
        and user.role
        and user.role.name == 'admin'
    )

def is_fired(user):
    return (
        user.status is not None
        and user.status.name == UserStatus.FIRED
    )

def normalize_value(value):
    if hasattr(value, 'tzinfo') and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value

def generate_excel(queryset, title, columns, headers):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title

        sheet.append(headers)

        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).font = Font(bold=True)

        for row in queryset:
            sheet.append([normalize_value(getattr(row, col)) for col in columns])

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        filename = title.lower().replace('-', '_')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

        workbook.save(response)

        return response
